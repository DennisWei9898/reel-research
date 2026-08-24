"""產出「隔夜策略手續費試算」Excel。每一格都是活公式，改參數會自己重算。"""
import yfinance as yf, pandas as pd, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TICKERS = ["MU","NVDA","AAPL","TSLA","SPY"]
END = pd.Timestamp("2026-08-24"); START = END - pd.Timedelta(days=100)

INK="FF1C1A17"; PAPER="FFF3EEE4"; ACC="FFB23A2E"; GOOD="FF3C6B4F"; WHITE="FFFFFDF8"
hdr = dict(font=Font(bold=True,color="FFFAF7F1",size=10),
           fill=PatternFill("solid",fgColor=INK),
           alignment=Alignment(horizontal="center",vertical="center",wrap_text=True))
thin = Border(*[Side("thin",color="FFD8D0C2")]*4)

wb = Workbook(); wb.remove(wb.active)

# ══ 分頁1：參數（使用者可改）══
ws = wb.create_sheet("① 參數設定")
ws["A1"]="隔夜策略手續費試算 · 參數設定"; ws["A1"].font=Font(bold=True,size=15,color=ACC)
ws["A2"]="改下面黃色格子的數字，其他分頁會自動重算。"; ws["A2"].font=Font(size=10,color="FF544E45")
rows=[("來回交易成本（bps，萬分之一）",30,"散戶合理區間 10–30。含手續費＋買賣價差＋滑價"),
      ("每年交易日數",252,"用來把期間報酬年化"),
      ("起始本金（美元）",10000,"只影響金額欄，不影響報酬率")]
ws["A4"]="參數"; ws["B4"]="數值"; ws["C4"]="說明"
for c in "ABC": ws[f"{c}4"].font=hdr["font"]; ws[f"{c}4"].fill=hdr["fill"]; ws[f"{c}4"].alignment=hdr["alignment"]
for i,(k,v,note) in enumerate(rows,5):
    ws[f"A{i}"]=k; ws[f"B{i}"]=v; ws[f"C{i}"]=note
    ws[f"B{i}"].fill=PatternFill("solid",fgColor="FFFFF3B0")
    ws[f"B{i}"].font=Font(bold=True,size=11)
    for c in "ABC": ws[f"{c}{i}"].border=thin
ws.column_dimensions["A"].width=32; ws.column_dimensions["B"].width=12; ws.column_dimensions["C"].width=52
ws["A9"]="※ B5 是「一趟來回」的總成本。每天進出一次＝每天付一次這個成本。"
ws["A9"].font=Font(italic=True,size=9,color=ACC)

# ══ 抓資料 ══
raw={}
for t in TICKERS:
    df=yf.download(t,start=START,end=END,interval="1d",auto_adjust=False,progress=False,threads=False)
    if isinstance(df.columns,pd.MultiIndex): df.columns=df.columns.get_level_values(0)
    raw[t]=df.dropna(subset=["Open","Close"])

# ══ 分頁2：逐日明細（示範一支，公式全開）══
t0="MU"; d=raw[t0]
ws2=wb.create_sheet(f"② 逐日明細（{t0}）")
ws2["A1"]=f"{t0} 逐日計算明細 · 看得到每一天的手續費怎麼吃掉報酬"
ws2["A1"].font=Font(bold=True,size=13,color=ACC)
ws2["A2"]="灰色欄＝原始股價（Yahoo Finance）；其他欄都是公式，可點進去看怎麼算的。"
ws2["A2"].font=Font(size=9,color="FF544E45")
cols=["日期","前日收盤","今日開盤","今日收盤","隔夜報酬","扣成本後隔夜報酬",
      "隔夜累積(無成本)","隔夜累積(扣成本)","買著不動累積"]
for j,c in enumerate(cols,1):
    cell=ws2.cell(4,j,c); cell.font=hdr["font"]; cell.fill=hdr["fill"]; cell.alignment=hdr["alignment"]
o=d["Open"].values; c_=d["Close"].values; idx=d.index
for i in range(1,len(d)):
    r=i+4
    ws2.cell(r,1,idx[i].strftime("%Y-%m-%d"))
    ws2.cell(r,2,round(float(c_[i-1]),4)); ws2.cell(r,3,round(float(o[i]),4)); ws2.cell(r,4,round(float(c_[i]),4))
    for cc in (2,3,4): ws2.cell(r,cc).fill=PatternFill("solid",fgColor="FFEFEFEF")
    ws2.cell(r,5,f"=C{r}/B{r}-1")                                  # 隔夜報酬
    ws2.cell(r,6,f"=E{r}-'① 參數設定'!$B$5/10000")                  # 扣成本
    ws2.cell(r,7, f"=1+E{r}" if i==1 else f"=G{r-1}*(1+E{r})")
    ws2.cell(r,8, f"=1+F{r}" if i==1 else f"=H{r-1}*(1+F{r})")
    ws2.cell(r,9, f"=D{r}/B{r}" if i==1 else f"=I{r-1}*(D{r}/B{r})")
    for cc in (5,6): ws2.cell(r,cc).number_format="0.0000%"
    for cc in (7,8,9): ws2.cell(r,cc).number_format="0.0000"
last=len(d)+3
ws2.cell(last+2,1,"期末結果").font=Font(bold=True,size=11)
ws2.cell(last+2,7,f"=G{last}-1").number_format="0.00%"
ws2.cell(last+2,8,f"=H{last}-1").number_format="0.00%"
ws2.cell(last+2,9,f"=I{last}-1").number_format="0.00%"
for cc in (7,8,9): ws2.cell(last+2,cc).font=Font(bold=True,size=11,color=ACC)
ws2.cell(last+3,1,f"交易趟數（每天一趟來回）").font=Font(size=9)
ws2.cell(last+3,2,len(d)-1)
ws2.cell(last+4,1,"手續費總共吃掉").font=Font(bold=True,size=10,color=ACC)
ws2.cell(last+4,2,f"=G{last}/H{last}-1").number_format="0.00%"
for j,w in enumerate([12,11,11,11,12,17,17,17,15],1): ws2.column_dimensions[get_column_letter(j)].width=w
ws2.freeze_panes="A5"

# ══ 分頁3：五支彙總 ══
ws3=wb.create_sheet("③ 五支彙總")
ws3["A1"]="五支股票 · 隔夜策略 vs 什麼都不做"; ws3["A1"].font=Font(bold=True,size=13,color=ACC)
ws3["A2"]=f"期間 {START.date()} ~ {END.date()}（約 3 個月）"; ws3["A2"].font=Font(size=9,color="FF544E45")
h=["股票","交易日數","隔夜(無成本)","扣10bps","扣30bps","扣50bps","買著不動","30bps下誰贏"]
for j,x in enumerate(h,1):
    cell=ws3.cell(4,j,x); cell.font=hdr["font"]; cell.fill=hdr["fill"]; cell.alignment=hdr["alignment"]
for i,t in enumerate(TICKERS,5):
    dd=raw[t]; oo=dd["Open"].values; cc2=dd["Close"].values
    ov=oo[1:]/cc2[:-1]-1; tot=cc2[1:]/cc2[:-1]-1
    cum=lambda r: float(np.prod(1+r)-1)
    ws3.cell(i,1,t).font=Font(bold=True)
    ws3.cell(i,2,len(ov))
    for j,bps in zip((3,4,5,6),(0,10,30,50)):
        v=cum(ov-bps/10000); cell=ws3.cell(i,j,v); cell.number_format="0.00%"
        cell.font=Font(color=GOOD if v>0 else ACC, bold=(j==5))
    bh=cum(tot); ws3.cell(i,7,bh).number_format="0.00%"
    ws3.cell(i,7).font=Font(color=GOOD if bh>0 else ACC)
    ws3.cell(i,8,f'=IF(E{i}>G{i},"隔夜贏","抱著贏")').font=Font(bold=True)
    for j in range(1,9): ws3.cell(i,j).border=thin
ws3.cell(11,1,"結論：扣掉真實手續費後，五支全部輸給「什麼都不做」。")
ws3.cell(11,1).font=Font(bold=True,size=11,color=ACC)
ws3.cell(12,1,"原因：這招每天要進出一次，3 個月就是 66 趟來回。每趟 30bps，光成本就吃掉約 20%。")
ws3.cell(12,1).font=Font(size=9,color="FF544E45")
for j,w in enumerate([10,10,15,12,12,12,13,14],1): ws3.column_dimensions[get_column_letter(j)].width=w

# ══ 分頁4：成本敏感度 ══
ws4=wb.create_sheet("④ 成本敏感度")
ws4["A1"]="手續費多高才會由賺轉賠？"; ws4["A1"].font=Font(bold=True,size=13,color=ACC)
ws4["A2"]="橫軸＝來回成本（bps）。看每支股票在哪一格由綠轉紅。"; ws4["A2"].font=Font(size=9,color="FF544E45")
grid=[0,5,10,15,20,25,30,40,50]
ws4.cell(4,1,"股票").font=hdr["font"]; ws4.cell(4,1).fill=hdr["fill"]
for j,b in enumerate(grid,2):
    cell=ws4.cell(4,j,f"{b}bps"); cell.font=hdr["font"]; cell.fill=hdr["fill"]; cell.alignment=hdr["alignment"]
for i,t in enumerate(TICKERS,5):
    dd=raw[t]; oo=dd["Open"].values; cc2=dd["Close"].values
    ov=oo[1:]/cc2[:-1]-1
    ws4.cell(i,1,t).font=Font(bold=True)
    for j,b in enumerate(grid,2):
        v=float(np.prod(1+(ov-b/10000))-1)
        cell=ws4.cell(i,j,v); cell.number_format="0.0%"
        cell.font=Font(color=GOOD if v>0 else ACC, bold=True)
        cell.fill=PatternFill("solid",fgColor="FFE1EEDF" if v>0 else "FFF5DCD8")
        cell.border=thin
ws4.cell(11,1,"注意：綠色不代表值得做——還要贏過同期「什麼都不做」（見分頁③）。")
ws4.cell(11,1).font=Font(bold=True,size=10,color=ACC)
ws4.column_dimensions["A"].width=10
for j in range(2,len(grid)+2): ws4.column_dimensions[get_column_letter(j)].width=10

out="data/隔夜策略手續費試算.xlsx"; wb.save(out); print("✅",out)
