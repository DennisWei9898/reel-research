#!/usr/bin/env python3
"""
照著一支 IG 教學 reel 做一次：=TEXTJOIN(" ",TRUE,G2:I2)
（@learn__excel「Textjoin function tricks」，26 秒、無旁白、無字幕）

這支腳本做三件事：
  1. 重建影片畫面上那張表（G=州、H=城市、I=語言），照影片打同一條公式
  2. 用 LibreOffice headless 真的算一次，把算出來的值抓回來跟影片畫面對照
  3. 額外測影片沒示範的部分：ignore_empty 給 TRUE 和 FALSE 到底差在哪、
     以及跟 CONCAT／& 的差別

執行：python3 textjoin_test.py     （需要 openpyxl 與已安裝的 LibreOffice）
"""
import csv, glob, os, subprocess, sys, shutil

try:
    from openpyxl import Workbook
except ImportError:
    sys.exit('請先安裝：pip install openpyxl')

SOFFICE = next((p for p in [
    '/Applications/LibreOffice.app/Contents/MacOS/soffice',
    shutil.which('soffice') or '',
    '/usr/bin/soffice',
] if p and os.path.exists(p)), None)
if not SOFFICE:
    sys.exit('找不到 LibreOffice。macOS: brew install --cask libreoffice')

# 影片畫面上前幾列的資料（G/H/I 三欄）
ROWS = [
    ('Gujarat', 'Surat', 'Gujarati'),
    ('Maharashtra', 'Pune', 'Marathi'),
    ('Rajasthan', 'Jaipur', 'Rajasthani'),
    ('Uttar Pradesh', 'Lucknow', 'Hindi'),
    ('Uttar Pradesh', 'Kanpur', 'Hindi'),
    ('Maharashtra', 'Nagpur', 'Marathi'),
    ('Madhya Pradesh', 'Indore', 'Hindi'),
    ('Maharashtra', 'Thane', 'Marathi'),
]
# 影片資料裡沒有的情況：中間那格是空的
# 🔴 xlsx 檔案格式的坑：TEXTJOIN／CONCAT／IFS 這些「後來才加的函式」
# 在檔案裡必須存成 _xlfn.TEXTJOIN。用程式寫檔（openpyxl）如果照人看到的名字寫，
# Excel 與 LibreOffice 都會回 #NAME? —— 這正是影片沒講的前置條件的檔案層版本。
XLFN = '=_xlfn.'

ROWS_WITH_GAP = [
    ('Gujarat', '', 'Gujarati'),
    ('', 'Pune', 'Marathi'),
    ('Bihar', 'Patna', ''),
]

def build(path):
    wb = Workbook(); ws = wb.active; ws.title = 'demo'
    ws['F1'], ws['G1'], ws['H1'], ws['I1'] = 'Status', 'Indian States', 'Indian Cities', 'Main Language'
    r = 2
    for st, city, lang in ROWS:
        ws.cell(r, 7, st); ws.cell(r, 8, city); ws.cell(r, 9, lang)
        ws.cell(r, 6, XLFN + f'TEXTJOIN(" ",TRUE,G{r}:I{r})')  # 影片教的那條
        r += 1
    gap_start = r
    for st, city, lang in ROWS_WITH_GAP:
        ws.cell(r, 7, st); ws.cell(r, 8, city); ws.cell(r, 9, lang)
        ws.cell(r, 6,  XLFN + f'TEXTJOIN(" ",TRUE,G{r}:I{r})')  # ignore_empty = TRUE
        ws.cell(r, 10, XLFN + f'TEXTJOIN(" ",FALSE,G{r}:I{r})') # ignore_empty = FALSE
        ws.cell(r, 11, '=_xlfn.CONCAT(G%d," ",H%d," ",I%d)'%(r,r,r))  # 舊做法
        ws.cell(r, 12, f'=G{r}&" "&H{r}&" "&I{r}')           # 更舊的做法
        r += 1
    # 分隔字元也可以是一整排（影片沒講）
    ws['N2'] = XLFN + 'TEXTJOIN(", ",TRUE,G2:G9)'
    wb.save(path)
    return gap_start, r - 1

def evaluate(xlsx):
    """用 LibreOffice 轉 CSV —— 轉檔會實際計算公式，等於真的跑了一次"""
    out = os.path.abspath('_calc')
    os.makedirs(out, exist_ok=True)
    subprocess.run([SOFFICE, '--headless', '--convert-to', 'csv:Text - txt - csv (StarCalc):44,34,76,1,,0,false,true,true',
                    '--outdir', out, xlsx],
                   check=True, capture_output=True, timeout=180)
    f = glob.glob(os.path.join(out, '*.csv'))[0]
    with open(f, encoding='utf-8') as fh:
        return list(csv.reader(fh))

def cell(grid, row, col):          # row/col 都是 1-based，對應 Excel 座標
    try: return grid[row - 1][col - 1]
    except IndexError: return ''

def main():
    xlsx = 'textjoin_demo.xlsx'
    gap_start, gap_end = build(xlsx)
    print(f'▶ 已建檔 {xlsx}，照影片打的公式：=TEXTJOIN(" ",TRUE,G2:I2)')
    grid = evaluate(xlsx)
    print('▶ LibreOffice 已實際計算完成\n')

    print('── ① 照影片做一次，結果對不對 ──')
    print(f'{"列":<4}{"影片公式算出來的 F 欄":<44}對照 G/H/I 三欄')
    for r in range(2, 2 + len(ROWS)):
        got = cell(grid, r, 6)
        src = ' / '.join(x for x in (cell(grid, r, 7), cell(grid, r, 8), cell(grid, r, 9)))
        print(f'{r:<4}{got:<44}{src}')
    expect = 'Gujarat Surat Gujarati'
    got2 = cell(grid, 2, 6)
    print(f'\n影片畫面上 F2 顯示：{expect}')
    print(f'我們算出來的 F2 ：{got2}')
    print('→ ' + ('✅ 完全一致，教學可重現' if got2 == expect else f'❌ 不一致：{got2!r}'))

    print('\n── ② 影片沒示範的：ignore_empty 給 TRUE 和 FALSE 差在哪 ──')
    print(f'{"列":<4}{"資料":<26}{"TRUE（影片用的）":<26}{"FALSE":<26}')
    same = 0
    for r in range(gap_start, gap_end + 1):
        src = '|'.join((cell(grid, r, 7), cell(grid, r, 8), cell(grid, r, 9)))
        t, f_ = cell(grid, r, 6), cell(grid, r, 10)
        if t == f_: same += 1
        print(f'{r:<4}{src:<26}{repr(t):<26}{repr(f_):<26}')
    print(f'\n影片示範用的 8 列資料完全沒有空格 → TRUE 與 FALSE 輸出必然相同。')
    print(f'只有在有空格時才看得出差別（上面 {gap_end-gap_start+1} 列有 {gap_end-gap_start+1-same} 列不同）。')

    print('\n── ③ 跟舊做法比 ──')
    r = gap_start
    print(f'資料：{cell(grid,r,7)!r} / {cell(grid,r,8)!r} / {cell(grid,r,9)!r}（中間是空的）')
    print(f'  TEXTJOIN(TRUE) : {cell(grid,r,6)!r}   ← 自動吃掉多餘的分隔字元')
    print(f'  CONCAT         : {cell(grid,r,11)!r}   ← 留下兩個空格')
    print(f'  用 & 串         : {cell(grid,r,12)!r}   ← 一樣留下兩個空格')

    print('\n── ④ 影片沒講的用法：分隔字元串一整欄 ──')
    print(f'  =TEXTJOIN(", ",TRUE,G2:G9) → {cell(grid,14,2) or cell(grid,2,14)!r}')

if __name__ == '__main__':
    main()
